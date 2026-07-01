"""Final solver for Problem 2.

Approach (deterministic, no Monte-Carlo):
1. For each candidate (size, z, dr, ty) over a coarse grid, build the
   staggered-ring layout centred on the tower (with tower allowed off-centre
   along the field's south-north axis).
2. Compute the per-mirror annual surrogate power (4 months x 3 times, full
   shadow/blocking).
3. Greedily keep mirrors in order of decreasing efficiency until either
   (a) we reach >=60 MW (TARGET), or (b) the field is exhausted.
4. If 60 MW is reachable, score by per-area thermal; otherwise score by raw
   annual MW.
5. Verify the top 3 designs at full fidelity (12 months x 5 times).
6. Write result2.xlsx and the figure set.
"""
import numpy as np, json, time
import engine as E
import openpyxl
import solve_p2_grid as G

TARGET_MW = 60.0


def select_and_eval(size, z, dr, ty, n_samp=3, full=True):
    """Build layout, rank, greedily select, run full annual."""
    W = H = size
    tower = (0.0, ty)
    pts = G.staggered_rings_in_field(W, dr, tower=tower)
    if len(pts) < 50:
        return None
    fld_full = E.Field(pts, W, H, z, tower=tower)
    pmw, peta = G.per_mirror_with_shadow(fld_full)
    order = np.argsort(-peta)
    cum = np.cumsum(pmw[order]) / 1000.0     # MW

    reach = cum[-1]
    if reach >= TARGET_MW:
        nsel = int(np.searchsorted(cum, TARGET_MW) + 1)
        rated = True
    else:
        nsel = len(pts)
        rated = False
    sel_pts = pts[order[:nsel]]
    fld = E.Field(sel_pts, W, H, z, tower=tower)
    if full:
        rows = fld.evaluate(do_shadow=True, n_samp=5)
        monthly, annual = E.monthly_annual(rows, fld.area.sum())
    else:
        monthly, annual = None, None
    return dict(size=size, z=z, dr=dr, ty=ty, gap=5.0,
                n=nsel, n_full=len(pts), area=float(fld.area.sum()),
                pts=sel_pts, W=W, H=H, rated=rated,
                surrogate_reach=float(reach),
                monthly=monthly, annual=annual)


def search_compact():
    """Smaller, finer grid centred on the most-promising region."""
    grid = []
    # densest 5m, 6m, 7m layouts at the minimum allowed dr (=W+5)
    grid += [(5, z, 10.0, ty) for z in [2.5, 3.5, 5.0]
                              for ty in [-150, -100, -50, 0, 50, 100]]
    grid += [(6, z, 11.0, ty) for z in [3.0, 4.0, 6.0]
                              for ty in [-150, -100, -50, 0, 50, 100]]
    grid += [(6, z, 12.0, ty) for z in [3.0, 4.0, 6.0]
                              for ty in [-150, -100, -50, 0]]
    grid += [(7, z, 12.0, ty) for z in [3.5, 5.0, 6.0]
                              for ty in [-150, -100, 0]]
    # also a few off-axis tower positions to see whether x-offset helps
    grid += [(5, 3.5, 10.0, ty) for ty in [-200]]
    rows = []
    t0 = time.time()
    print(f"{'design':40s} {'n':>5} {'reach':>6} {'P':>6} {'ppa':>7}")
    for size, z, dr, ty in grid:
        r = select_and_eval(size, z, dr, ty)
        if r is None:
            continue
        ann = r['annual']
        tag = f"sz={size} z={z} dr={dr:.0f} ty={ty:+.0f}"
        print(f"{tag:40s} {r['n']:>5} {r['surrogate_reach']:>6.2f} "
              f"{ann['power']:>6.2f} {ann['ppa']:>7.4f}  "
              f"{'RATED' if r['rated'] else 'best-effort'}")
        rows.append(r)
    print(f"elapsed {time.time()-t0:.1f}s")
    return rows


def main():
    rows = search_compact()
    rated = [r for r in rows if r['rated']]
    if rated:
        pool = rated
        print("\nChoosing best per-area among rated designs:")
    else:
        pool = rows
        print("\nNo design reaches 60 MW; picking highest annual MW:")
    pool.sort(key=lambda r: -r['annual']['ppa'])
    for r in pool[:5]:
        ann = r['annual']
        print(f"  sz={r['size']} z={r['z']} dr={r['dr']:.0f} ty={r['ty']:+.0f}: "
              f"P={ann['power']:.3f}  ppa={ann['ppa']:.4f}  eta={ann['eta']:.4f}")
    if not rated:
        # if not rated, pick highest power instead
        pool.sort(key=lambda r: -r['annual']['power'])

    best = pool[0]
    out = {
        'design': dict(size=best['size'], z=best['z'], dr=best['dr'],
                       ty=best['ty'], gap=best['gap'], n=best['n'],
                       area=best['area'], W=best['W'], H=best['H'],
                       rated=best['rated'], reach=best['surrogate_reach']),
        'monthly': {str(m): best['monthly'][m] for m in best['monthly']},
        'annual': best['annual'],
        'candidates': [{
            'size': r['size'], 'z': r['z'], 'dr': r['dr'], 'ty': r['ty'],
            'n': r['n'], 'area': r['area'], 'rated': r['rated'],
            'power': r['annual']['power'], 'ppa': r['annual']['ppa'],
            'eta': r['annual']['eta']} for r in pool],
    }
    with open('../results/problem2.json', 'w') as f:
        json.dump(out, f, indent=2, default=float)
    np.save('p2_final_pts.npy', best['pts'])

    # ---- write result2.xlsx (per the problem's "template" format) ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '设计参数'
    headers = ['吸收塔位置x(m)', '吸收塔位置y(m)',
               '定日镜宽度(m)', '定日镜高度(m)',
               '定日镜安装高度(m)', '定日镜数目',
               '定日镜总面积(m^2)']
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    ws.cell(row=2, column=1, value=0.0)
    ws.cell(row=2, column=2, value=float(best['ty']))
    ws.cell(row=2, column=3, value=float(best['W']))
    ws.cell(row=2, column=4, value=float(best['H']))
    ws.cell(row=2, column=5, value=float(best['z']))
    ws.cell(row=2, column=6, value=int(best['n']))
    ws.cell(row=2, column=7, value=float(best['area']))

    ws2 = wb.create_sheet('定日镜位置')
    ws2.cell(row=1, column=1, value='x坐标(m)')
    ws2.cell(row=1, column=2, value='y坐标(m)')
    for i, (x, y) in enumerate(best['pts'], start=2):
        ws2.cell(row=i, column=1, value=float(x))
        ws2.cell(row=i, column=2, value=float(y))
    wb.save('../results/result2.xlsx')
    print("\nwrote ../results/problem2.json and ../results/result2.xlsx")
    print('Best design:', json.dumps(out['design'], indent=2, default=float))
    print('Annual:', json.dumps(out['annual'], indent=2, default=float))


if __name__ == '__main__':
    main()
