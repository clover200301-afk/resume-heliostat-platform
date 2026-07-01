"""Re-pick best-by-ppa from problem2.json candidates and rebuild
   result2.xlsx + full evaluation of THAT design."""
import json, numpy as np, time
import engine as E
import solve_p2_grid as G
import openpyxl

# from previous run
d = json.load(open('../results/problem2.json'))
cs = sorted(d['candidates'], key=lambda c: -c['ppa'])
print('Top 3 by ppa:')
for c in cs[:3]:
    print(c)

best_cand = cs[0]
size, z, dr, ty = best_cand['size'], best_cand['z'], best_cand['dr'], best_cand['ty']
W = H = size
tower = (0.0, ty)

pts = G.staggered_rings_in_field(W, dr, tower=tower)
print(f"\nBuilding {size}x{size} layout, z={z}, dr={dr}, ty={ty}: n_full={len(pts)}")
fld_full = E.Field(pts, W, H, z, tower=tower)
pmw, peta = G.per_mirror_with_shadow(fld_full)
order = np.argsort(-peta)
nsel = best_cand['n']
sel_pts = pts[order[:nsel]]
print(f"keeping top {nsel}")

fld = E.Field(sel_pts, W, H, z, tower=tower)
t0 = time.time()
rows = fld.evaluate(do_shadow=True, n_samp=5)
print(f"full eval ({time.time()-t0:.1f}s)")
monthly, annual = E.monthly_annual(rows, fld.area.sum())
print('Annual:', json.dumps(annual, indent=2, default=float))

# update problem2.json
out = {
    'design': dict(size=size, z=z, dr=dr, ty=ty, gap=5.0,
                   n=nsel, area=float(fld.area.sum()), W=W, H=H,
                   reach=float(np.cumsum(pmw[order])[-1] / 1000.0),
                   rated=False),
    'monthly': {str(m): monthly[m] for m in monthly},
    'annual': annual,
    'candidates': d['candidates'],
}
with open('../results/problem2.json', 'w') as f:
    json.dump(out, f, indent=2, default=float)
np.save('p2_final_pts.npy', sel_pts)

# write result2.xlsx
wb = openpyxl.Workbook()
ws = wb.active; ws.title = '设计参数'
for i, h in enumerate(['吸收塔位置x(m)', '吸收塔位置y(m)',
                       '定日镜宽度(m)', '定日镜高度(m)',
                       '定日镜安装高度(m)', '定日镜数目',
                       '定日镜总面积(m^2)'], 1):
    ws.cell(row=1, column=i, value=h)
ws.cell(row=2, column=1, value=0.0)
ws.cell(row=2, column=2, value=float(ty))
ws.cell(row=2, column=3, value=float(W))
ws.cell(row=2, column=4, value=float(H))
ws.cell(row=2, column=5, value=float(z))
ws.cell(row=2, column=6, value=int(nsel))
ws.cell(row=2, column=7, value=float(fld.area.sum()))

ws2 = wb.create_sheet('定日镜位置')
ws2.cell(row=1, column=1, value='x坐标(m)')
ws2.cell(row=1, column=2, value='y坐标(m)')
for i, (x, y) in enumerate(sel_pts, start=2):
    ws2.cell(row=i, column=1, value=float(x))
    ws2.cell(row=i, column=2, value=float(y))
wb.save('../results/result2.xlsx')
print("\nwrote ../results/problem2.json and ../results/result2.xlsx")
