"""Problem 3 solver: zonal mirror sizing.

Strategy:
  * Divide the field around the tower into K concentric zones by radius.
  * In each zone use the staggered-ring layout with that zone's (size, z, dr).
  * Optimize per-zone parameters; jointly select mirrors with best annual
    surrogate until annual >= 60 MW; report per-area.

The reasoning: smaller mirrors near the tower keep shadow/blocking low while
gaining little extra area; larger mirrors far from the tower hold more area
per packed slot. Allowing variable install height also lets distant mirrors
sit higher (avoiding blocking from inner rings).
"""
import numpy as np, json, time, itertools
import engine as E
import openpyxl

R_FIELD = 350.0
R_EXCL_TOWER = 100.0
TARGET_MW = 60.0


def zone_layout(zones, tower=(0.0, 0.0), gap=5.0):
    """zones = [(r_outer, size, z, dr), ...] sorted by r_outer ascending.
    First zone starts at R_EXCL_TOWER."""
    tx, ty = tower
    r_inner = R_EXCL_TOWER
    all_pts = []
    sizes = []
    zs = []
    for (r_outer, size, z, dr) in zones:
        # rings between r_inner and r_outer with pitch dr
        min_sep = size + gap
        # start half-pitch above r_inner
        r = r_inner + dr * 0.5
        ring = 0
        while r <= r_outer + 1e-6:
            n = max(1, int(np.floor(2 * np.pi * r / min_sep)))
            dphi = 2 * np.pi / n
            offset = (ring % 2) * dphi * 0.5
            for k in range(n):
                phi = offset + k * dphi
                x = tx + r * np.cos(phi); y = ty + r * np.sin(phi)
                if np.hypot(x, y) <= R_FIELD - 1e-6:
                    all_pts.append((x, y))
                    sizes.append(size)
                    zs.append(z)
            r += dr
            ring += 1
        r_inner = r_outer
    pts = np.array(all_pts) if all_pts else np.zeros((0, 2))
    return pts, np.array(sizes), np.array(zs)


def per_mirror_eval(pts, sizes, zs, tower, n_samp=3,
                    months=(3, 6, 9, 12), times=(9.0, 12.0, 15.0)):
    fld = E.Field(pts, sizes, sizes, zs, tower=tower)
    at = fld.atmospheric_eff()
    acc_p = np.zeros(fld.N); acc_e = np.zeros(fld.N)
    k = 0
    for m in months:
        D = E.day_number(m)
        for ST in times:
            s, alt = E.sun_vector(D, ST)
            if alt <= 0:
                continue
            I = E.dni(alt)
            cosv, n = fld.cosine_eff(s)
            trunc = fld.truncation_eff(s, n)
            sb = fld.shadow_block_eff(s, n, n_samp=n_samp)
            eta = sb * cosv * at * trunc * E.ETA_REF
            acc_p += I * fld.area * eta
            acc_e += eta
            k += 1
    return acc_p / k, acc_e / k, fld


def evaluate_zones(zones, tower=(0.0, -50.0), full=False):
    pts, sizes, zs = zone_layout(zones, tower=tower)
    if len(pts) < 50:
        return None
    pmw, peta, fld_full = per_mirror_eval(pts, sizes, zs, tower)
    order = np.argsort(-peta)
    cum = np.cumsum(pmw[order]) / 1000.0
    reach = cum[-1]
    if reach >= TARGET_MW:
        nsel = int(np.searchsorted(cum, TARGET_MW) + 1)
        rated = True
    else:
        nsel = len(pts)
        rated = False
    sel = order[:nsel]
    sel_pts = pts[sel]
    sel_sizes = sizes[sel]
    sel_zs = zs[sel]
    sel_area = sel_sizes * sel_sizes
    total_area = sel_area.sum()
    monthly = annual = None
    if full:
        fld = E.Field(sel_pts, sel_sizes, sel_sizes, sel_zs, tower=tower)
        rows = fld.evaluate(do_shadow=True, n_samp=5)
        monthly, annual = E.monthly_annual(rows, total_area)
    surrogate_power = cum[nsel - 1]
    surrogate_ppa = surrogate_power * 1000.0 / total_area
    return dict(zones=zones, tower=tower, rated=rated,
                reach=float(reach), n=int(nsel),
                area=float(total_area),
                surrogate_power=float(surrogate_power),
                surrogate_ppa=float(surrogate_ppa),
                pts=sel_pts, sizes=sel_sizes, zs=sel_zs,
                monthly=monthly, annual=annual)


def grid_search():
    """Try a small set of plausible zonal designs."""
    designs = []
    # uniform baselines (P2's optimal: 5x5/z2.5/dr10)
    for sz, dr in [(4, 9.0), (5, 10.0), (6, 11.0), (7, 12.0)]:
        designs.append([(350.0, sz, max(2, sz/2), dr)])
    # 2 zones with finer combinations
    for r1 in [150, 180, 220, 260, 300]:
        for sin in [3, 4, 5]:
            for sout in [5, 6, 7]:
                if sout <= sin: continue
                designs.append([(r1, sin, max(2, sin/2), sin + 5.0),
                                (350.0, sout, max(2, sout/2), sout + 5.0)])
    # 3 zones
    for r1, r2 in [(160, 240), (170, 260), (200, 280), (180, 270)]:
        for sa, sb_, sc in [(3, 5, 7), (3, 4, 6), (4, 5, 6), (4, 5, 7),
                            (3, 5, 6), (4, 6, 7)]:
            designs.append([(r1, sa, max(2, sa/2), sa + 5.0),
                            (r2, sb_, max(2, sb_/2), sb_ + 5.0),
                            (350.0, sc, max(2, sc/2), sc + 5.0)])
    # vary install height in remote zone (slightly higher mirrors)
    for r1, r2 in [(180, 270)]:
        for sa, sb_, sc in [(4, 5, 6), (4, 5, 7)]:
            designs.append([(r1, sa, max(2, sa/2), sa + 5.0),
                            (r2, sb_, max(2, sb_/2), sb_ + 5.0),
                            (350.0, sc, min(6, sc), sc + 5.0)])
    # towers — be less restrictive
    towers = [(0.0, 0.0), (0.0, -50.0), (0.0, -100.0)]
    results = []
    for design in designs:
        for tw in towers:
            r = evaluate_zones(design, tower=tw, full=False)
            if r is None:
                continue
            results.append(r)
            tag = '+'.join(f"<{z[0]:.0f}:{z[1]}x{z[1]}@z{z[2]}/dr{z[3]:.0f}>" for z in design)
            print(f"tw=({tw[0]:+.0f},{tw[1]:+.0f}) {tag[:80]}  "
                  f"reach={r['reach']:.2f}  n={r['n']}  ppa={r['surrogate_ppa']:.4f}  "
                  f"{'RATED' if r['rated'] else 'best-eff'}")
    return results


def main():
    print("--- surrogate search ---")
    t0 = time.time()
    res = grid_search()
    print(f"surrogate search done in {time.time()-t0:.1f}s")
    # If 60 MW is unattainable, the problem still asks to reach the rated
    # value as closely as possible while maximising ppa.  We score by
    # weighted criterion: reach × ppa  (a Pareto-front proxy that prefers
    # designs which are both close to 60 MW and have high per-area density).
    def score(r):
        # cap reach at TARGET so we don't over-reward designs above 60 MW
        return min(r['reach'], TARGET_MW) * r['surrogate_ppa']
    res.sort(key=lambda r: -score(r))
    print("\nTop 8 surrogate by reach × ppa:")
    for r in res[:8]:
        print(f"  n={r['n']} area={r['area']:.0f} ppa={r['surrogate_ppa']:.4f} "
              f"reach={r['reach']:.2f} score={score(r):.2f} tower={r['tower']} "
              f"zones={r['zones']}")

    # full-evaluate the top 4 candidates
    print("\n--- full evaluation of top 4 ---")
    full_results = []
    for r in res[:4]:
        rf = evaluate_zones(r['zones'], tower=r['tower'], full=True)
        ann = rf['annual']
        print(f"  zones={r['zones']} tw={r['tower']} -> "
              f"P={ann['power']:.3f} MW  ppa={ann['ppa']:.4f}  eta={ann['eta']:.4f}")
        full_results.append(rf)
    # final selection: balance P close to 60 MW with high ppa
    full_results.sort(key=lambda r: -(min(r['annual']['power'], TARGET_MW)
                                       * r['annual']['ppa']))
    best = full_results[0]
    ann = best['annual']
    print(f"\nBEST: P={ann['power']:.3f}  ppa={ann['ppa']:.4f}  eta={ann['eta']:.4f}")

    # save
    out = {
        'design': dict(zones=best['zones'], tower=best['tower'],
                       n=best['n'], area=best['area'],
                       rated=best['rated']),
        'monthly': {str(m): best['monthly'][m] for m in best['monthly']},
        'annual': best['annual'],
    }
    with open('../results/problem3.json', 'w') as f:
        json.dump(out, f, indent=2, default=float)
    np.save('p3_final_pts.npy', best['pts'])
    np.save('p3_final_sizes.npy', best['sizes'])
    np.save('p3_final_zs.npy', best['zs'])

    # write result3.xlsx
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '设计参数'
    ws['A1'] = '吸收塔位置x(m)'; ws['B1'] = '吸收塔位置y(m)'
    ws['C1'] = '定日镜数目'; ws['D1'] = '定日镜总面积(m^2)'
    ws['A2'] = float(best['tower'][0]); ws['B2'] = float(best['tower'][1])
    ws['C2'] = int(best['n']); ws['D2'] = float(best['area'])

    ws2 = wb.create_sheet('定日镜')
    for i, h in enumerate(['x(m)', 'y(m)', '宽度(m)', '高度(m)', '安装高度(m)'], 1):
        ws2.cell(row=1, column=i, value=h)
    for i, ((x, y), s, z) in enumerate(zip(best['pts'], best['sizes'], best['zs']), start=2):
        ws2.cell(row=i, column=1, value=float(x))
        ws2.cell(row=i, column=2, value=float(y))
        ws2.cell(row=i, column=3, value=float(s))
        ws2.cell(row=i, column=4, value=float(s))
        ws2.cell(row=i, column=5, value=float(z))
    wb.save('../results/result3.xlsx')
    print("wrote ../results/problem3.json and ../results/result3.xlsx")


if __name__ == '__main__':
    main()
