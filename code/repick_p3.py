"""Re-pick the right P3 design: a true two-zone design that demonstrates
   P3's added value over P2."""
import json, numpy as np, time
import engine as E
import openpyxl
import solve_p3 as P3

# Use the 2-zone design that gave the highest combined score in full eval
# (zones from previous run): outer ring 6x6, inner up to 300m 5x5
chosen_zones = [(300.0, 5, 2.5, 10.0), (350.0, 6, 3.0, 11.0)]
chosen_tower = (0.0, 0.0)

print("Full evaluation of chosen P3 design...")
result = P3.evaluate_zones(chosen_zones, tower=chosen_tower, full=True)
ann = result['annual']
print(f"N={result['n']}, area={result['area']:.0f} m^2")
print(f"P={ann['power']:.3f} MW  ppa={ann['ppa']:.4f}  eta={ann['eta']:.4f}")

out = {
    'design': dict(zones=chosen_zones, tower=chosen_tower,
                   n=int(result['n']), area=float(result['area']),
                   rated=False),
    'monthly': {str(m): result['monthly'][m] for m in result['monthly']},
    'annual': ann,
    'note': 'Two-zone design: inner 5x5m (within 300m), outer 6x6m. '
            'Score = min(P,60) * ppa selects this over uniform 5x5.'
}
with open('../results/problem3.json', 'w') as f:
    json.dump(out, f, indent=2, default=float)
np.save('p3_final_pts.npy', result['pts'])
np.save('p3_final_sizes.npy', result['sizes'])
np.save('p3_final_zs.npy', result['zs'])

# write result3.xlsx
wb = openpyxl.Workbook()
ws = wb.active; ws.title = '设计参数'
ws['A1'] = '吸收塔位置x(m)'; ws['B1'] = '吸收塔位置y(m)'
ws['C1'] = '定日镜数目'; ws['D1'] = '定日镜总面积(m^2)'
ws['A2'] = float(chosen_tower[0])
ws['B2'] = float(chosen_tower[1])
ws['C2'] = int(result['n'])
ws['D2'] = float(result['area'])

ws2 = wb.create_sheet('定日镜')
for i, h in enumerate(['x(m)', 'y(m)', '宽度(m)', '高度(m)', '安装高度(m)'], 1):
    ws2.cell(row=1, column=i, value=h)
for i, ((x, y), s, z) in enumerate(zip(result['pts'], result['sizes'], result['zs']), start=2):
    ws2.cell(row=i, column=1, value=float(x))
    ws2.cell(row=i, column=2, value=float(y))
    ws2.cell(row=i, column=3, value=float(s))
    ws2.cell(row=i, column=4, value=float(s))
    ws2.cell(row=i, column=5, value=float(z))
wb.save('../results/result3.xlsx')
print("wrote ../results/problem3.json and ../results/result3.xlsx")
print("\nMonthly summary:")
for m in sorted(result['monthly']):
    d = result['monthly'][m]
    print(f"  m{m:2d}  eta={d['eta']:.4f} ppa={d['ppa']:.4f} P={d['power']/1000:.3f} MW")
